from __future__ import annotations

import math
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont




SEVERITY_COLORS: Dict[str, Optional[str]] = {
    "very low": "00B050",
    "low": "00B050",
    "medium": "FFFF00",
    "high": "F4B183",
    "very high": "FF0000",
    "critical": "FF0000",
    "unknown": None,
    "unknow": None,
}


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E2F3")
HEADER_FONT = Font(bold=True, size=13)

THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)



# Utility functions


def normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def find_column_index(ws, column_name: str) -> Optional[int]:
    for col in range(1, ws.max_column + 1):
        if normalize_text(ws.cell(1, col).value).lower() == column_name.lower():
            return col
    return None


def delete_column_if_needed(ws, column_name: str, should_delete: bool) -> None:
    idx = find_column_index(ws, column_name)
    if idx and should_delete:
        ws.delete_cols(idx, 1)


def all_values_are_na(values: List[str]) -> bool:
    cleaned = [v for v in values if v != ""]
    if not cleaned:
        return False
    return all(v.upper() == "N/A" for v in cleaned)


def all_values_empty(values: List[str]) -> bool:
    return all(v == "" for v in values)



# Formatting


def apply_basic_formatting(ws):

    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col)
        header.fill = HEADER_FILL
        header.font = HEADER_FONT
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header.border = THIN_BORDER

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER


def color_severity_column(ws):

    idx = find_column_index(ws, "Severity")
    if not idx:
        return

    for row in range(2, ws.max_row + 1):

        cell = ws.cell(row, idx)
        text = normalize_text(cell.value).lower()

        fill = None

        if "very low" in text:
            fill = SEVERITY_COLORS["very low"]

        elif "very high" in text:
            fill = SEVERITY_COLORS["very high"]

        elif "critical" in text:
            fill = SEVERITY_COLORS["critical"]

        elif text in ["unknown", "unknow"]:
            fill = None

        elif "medium" in text:
            fill = SEVERITY_COLORS["medium"]

        elif "high" in text:
            fill = SEVERITY_COLORS["high"]

        elif "low" in text:
            fill = SEVERITY_COLORS["low"]

        if fill:
            cell.fill = PatternFill(fill_type="solid", fgColor=fill)



# Layout optimisation


def optimize_layout(ws):

    preferred_widths = {
        "Created": 18,
        "Position in current video": 16,
        "Position from first video": 16,
        "Position from the start (m)": 12,
        "Distance from the previous manhole (m)": 18,
        "Code": 8,
        "Characteristic 1": 12,
        "Characteristic 2": 12,
        "Observation type": 34,
        "Clockface references": 12,
        "Continuing defect": 12,
        "End of": 12,
        "Observation step": 12,
        "Note": 24,
        "Severity": 12,
        "Longitude": 12,
        "Latitude": 12,
    }

    min_width = 8
    max_width_default = 18

    for col in range(1, ws.max_column + 1):

        header = normalize_text(ws.cell(1, col).value)

        if header in preferred_widths:
            ws.column_dimensions[get_column_letter(col)].width = preferred_widths[header]
            continue

        max_len = len(header)

        for row in range(2, ws.max_row + 1):

            value = normalize_text(ws.cell(row, col).value)

            line_max = max((len(part) for part in value.splitlines()), default=0)

            max_len = max(max_len, line_max)

        estimated = min(max(max_len + 2, min_width), max_width_default)

        ws.column_dimensions[get_column_letter(col)].width = estimated
# Render excel to image code

def rgb_from_openpyxl_color(cell_fill) -> Tuple[int, int, int]:
    try:
        if cell_fill is None or cell_fill.fill_type != "solid":
            return (255, 255, 255)

        fg = cell_fill.fgColor
        color_type = getattr(fg, "type", None)

        if color_type != "rgb":
            return (255, 255, 255)

        color = getattr(fg, "rgb", None)
        if not color:
            return (255, 255, 255)

        color = color.upper()
        if color in {"00000000", "000000", "000000FF"}:
            return (255, 255, 255)

        if len(color) == 8:
            color = color[2:]
        if len(color) != 6:
            return (255, 255, 255)

        return tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
    except Exception:
        return (255, 255, 255)


def load_font(size: int, bold: bool = False):
    candidates = []
    if bold:
        candidates.extend([
            "times.ttf",
            "times.ttf",
        ])
    else:
        candidates.extend([
            "times.ttf",
            "times.ttf",
        ])

    for path in candidates:
        if Path(path).exists():
            return ImageFont.truetype(path, size=size)

    return ImageFont.load_default()

def wrap_text_for_pixels(text: str, font, max_width_px: int) -> List[str]:
    text = text or ""
    paragraphs = text.splitlines() or [""]
    lines: List[str] = []

    dummy_img = Image.new("RGB", (10, 10))
    draw = ImageDraw.Draw(dummy_img)

    for paragraph in paragraphs:
        words = paragraph.split(" ")
        if not words:
            lines.append("")
            continue

        current = words[0]
        for word in words[1:]:
            test = f"{current} {word}".strip()
            bbox = draw.textbbox((0, 0), test, font=font)
            test_width = bbox[2] - bbox[0]

            if test_width <= max_width_px:
                current = test
            else:
                lines.append(current)
                current = word

        lines.append(current)

    return lines or [""]


def get_text_block_size(lines: List[str], font, draw) -> Tuple[int, int, List[int]]:
    line_heights = []
    max_width = 0

    for line in lines:
        bbox = draw.textbbox((0, 0), line or " ", font=font)
        w = bbox[2] - bbox[0]
        h = bbox[3] - bbox[1]
        max_width = max(max_width, w)
        line_heights.append(h)

    total_height = sum(line_heights)
    if len(line_heights) > 1:
        total_height += 2 * (len(line_heights) - 1)

    return max_width, total_height, line_heights


def draw_cell_text(
    draw,
    cell,
    x: int,
    y: int,
    cell_w: int,
    cell_h: int,
    font,
    padding_x: int = 8,
    padding_y: int = 6,
):
    text = normalize_text(cell.value)
    lines = wrap_text_for_pixels(text, font, max_width_px=max(10, cell_w - 2 * padding_x))

    text_block_w, text_block_h, line_heights = get_text_block_size(lines, font, draw)

    alignment = cell.alignment
    horizontal = (alignment.horizontal or "").lower()
    vertical = (alignment.vertical or "").lower()

    # Horizontal position
    if horizontal == "center":
        text_x = x + max(padding_x, (cell_w - text_block_w) // 2)
    elif horizontal == "right":
        text_x = x + max(padding_x, cell_w - text_block_w - padding_x)
    else:
        text_x = x + padding_x

    # Vertical position
    if vertical == "center":
        text_y = y + max(padding_y, (cell_h - text_block_h) // 2)
    elif vertical == "bottom":
        text_y = y + max(padding_y, cell_h - text_block_h - padding_y)
    else:
        text_y = y + padding_y

    # Draw each line
    current_y = text_y
    for i, line in enumerate(lines):
        line_bbox = draw.textbbox((0, 0), line or " ", font=font)
        line_w = line_bbox[2] - line_bbox[0]

        if horizontal == "center":
            line_x = x + max(padding_x, (cell_w - line_w) // 2)
        elif horizontal == "right":
            line_x = x + max(padding_x, cell_w - line_w - padding_x)
        else:
            line_x = text_x

        draw.text((line_x, current_y), line, fill="black", font=font)
        current_y += line_heights[i] + 2


def render_sheet_to_png(ws, output_png: str) -> None:
    header_font = load_font(18, bold=True)
    body_font = load_font(17, bold=False)

    # Convert Excel widths/heights to pixels
    col_widths_px: List[int] = []
    for col in range(1, ws.max_column + 1):
        width = ws.column_dimensions[get_column_letter(col)].width or 10
        col_widths_px.append(int(width * 8 + 12))

    row_heights_px: List[int] = []
    for row in range(1, ws.max_row + 1):
        height = ws.row_dimensions[row].height or 20
        row_heights_px.append(int(height * 2.5))

    img_width = sum(col_widths_px) + 1
    img_height = sum(row_heights_px) + 1

    img = Image.new("RGB", (img_width, img_height), "white")
    draw = ImageDraw.Draw(img)

    y = 0
    for r in range(1, ws.max_row + 1):
        x = 0
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell_w = col_widths_px[c - 1]
            cell_h = row_heights_px[r - 1]

            fill_rgb = rgb_from_openpyxl_color(cell.fill)

            # Cell background + border
            draw.rectangle(
                [x, y, x + cell_w, y + cell_h],
                fill=fill_rgb,
                outline="black",
                width=1
            )

            active_font = header_font if r == 1 else body_font

            draw_cell_text(
                draw=draw,
                cell=cell,
                x=x,
                y=y,
                cell_w=cell_w,
                cell_h=cell_h,
                font=active_font,
                padding_x=8,
                padding_y=6,
            )

            x += cell_w
        y += row_heights_px[r - 1]

    img.save(output_png)



# Main processing function


def process_excel(
    input_xlsx: str,
    output_xlsx: str,
    output_png: str,
    sheet_name: Optional[str] = None
):

    wb = load_workbook(input_xlsx)

    ws = wb[sheet_name] if sheet_name else wb.active


    # delete Video column
    delete_column_if_needed(ws, "Video", True)


    # Observation step
    idx = find_column_index(ws, "Observation step")

    if idx:
        values = [normalize_text(ws.cell(r, idx).value) for r in range(2, ws.max_row + 1)]
        delete_column_if_needed(ws, "Observation step", all_values_are_na(values))


    # Note column
    idx = find_column_index(ws, "Note")

    if idx:
        values = [normalize_text(ws.cell(r, idx).value) for r in range(2, ws.max_row + 1)]
        delete_column_if_needed(ws, "Note", all_values_empty(values))


    apply_basic_formatting(ws)

    color_severity_column(ws)

    optimize_layout(ws)

    wb.save(output_xlsx)
    render_sheet_to_png(ws, output_png)


    return output_xlsx, output_png
